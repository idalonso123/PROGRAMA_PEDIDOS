#!/usr/bin/env python3
"""
Motor de Cálculo ABC+D para Gestión de Inventarios - VERSIÓN V2
Vivero Aranjuez - Sistema de Clasificación ABC+D por Períodos

Este script combina:
- Cálculo de clasificación ABC+D por períodos definidos
- Aplicación de formatos Excel profesionales
- Soporte para nuevo formato de archivos del ERP (SPA_*.xlsx)
- Procesamiento de múltiples secciones
- Envío automático de emails a los encargados de cada sección

MODO DE USO:
- Con parámetro --periodo: Procesa el período especificado (1-4)
- Con parámetro --seccion: Procesa solo la sección especificada

Ejecutar: 
    python clasificacionABC.py --periodo 1              # Procesa Período 1 (enero-febrero)
    python clasificacionABC.py --periodo 2              # Procesa Período 2 (marzo-mayo)
    python clasificacionABC.py --periodo 3              # Procesa Período 3 (junio-agosto)
    python clasificacionABC.py --periodo 4              # Procesa Período 4 (septiembre-diciembre)
    python clasificacionABC.py --periodo 1 --seccion vivero  # Solo sección vivero del período 1

Los datos se leen de archivos con formato ERP:
- SPA_Compras.xlsx: Datos de compras del año
- SPA_Ventas.xlsx: Datos de ventas del año
- SPA_Stock.xlsx: Datos de stock actual
- SPA_Coste.xlsx: Costes unitarios de artículos (para calcular beneficio real)

Al generar cada archivo de clasificación, se envía automáticamente un email
al encargado de la sección con el archivo adjunto.

Períodos definidos:
- Período 1: 1 enero - 28 febrero (59 días)
- Período 2: 1 marzo - 31 mayo (92 días)
- Período 3: 1 junio - 31 agosto (92 días)
- Período 4: 1 septiembre - 31 diciembre (122 días)
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
import sys
import argparse
import warnings
import smtplib
import ssl
import os
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from pathlib import Path

# Importar el módulo de configuración centralizada
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))
from config_loader import (
    get_abc_config,
    get_secciones_config,
    get_encargados_config,
    get_smtp_config,
    get_rotaciones_familia,
    get_iva_familia,
    get_iva_subfamilia,
    calcular_periodo_ventas
)
from file_finder import find_latest_file

warnings.filterwarnings('ignore')

# ============================================================================
# DEFINICIÓN DE PERÍODOS (según manual del sistema V2)
# ============================================================================

DEFINICION_PERIODOS = {
    1: {
        'nombre': 'Periodo_1',
        'descripcion': 'Enero - Febrero',
        'fecha_inicio': '01/01',
        'fecha_fin': '28/02',
        'dias': 59
    },
    2: {
        'nombre': 'Periodo_2',
        'descripcion': 'Marzo - Mayo',
        'fecha_inicio': '01/03',
        'fecha_fin': '31/05',
        'dias': 92
    },
    3: {
        'nombre': 'Periodo_3',
        'descripcion': 'Junio - Agosto',
        'fecha_inicio': '01/06',
        'fecha_fin': '31/08',
        'dias': 92
    },
    4: {
        'nombre': 'Periodo_4',
        'descripcion': 'Septiembre - Diciembre',
        'fecha_inicio': '01/09',
        'fecha_fin': '31/12',
        'dias': 122
    }
}

# ============================================================================
# FUNCIONES DE NORMALIZACIÓN PARA BÚSQUEDA FLEXIBLE DE COLUMNAS
# ============================================================================

import unicodedata

def normalizar_texto(texto):
    """
    Normaliza un texto eliminando acentos y convirtiéndolo a minúsculas.
    
    Args:
        texto: String a normalizar
    
    Returns:
        str: Texto normalizado (sin acentos y en minúsculas)
    """
    if pd.isna(texto):
        return ''
    # Eliminar acentos usando unicodedata
    texto_normalizado = unicodedata.normalize('NFD', str(texto))
    texto_sin_acentos = ''.join(c for c in texto_normalizado if unicodedata.category(c) != 'Mn')
    return texto_sin_acentos.lower().strip()

def buscar_columna_normalizada(df, nombre_columna):
    """
    Busca una columna en el DataFrame ignorando mayúsculas/minúsculas y acentos.
    
    Args:
        df: DataFrame donde buscar
        nombre_columna: Nombre de la columna a buscar
    
    Returns:
        str: Nombre de la columna encontrada o None si no existe
    """
    nombre_normalizado = normalizar_texto(nombre_columna)
    
    for col in df.columns:
        if normalizar_texto(col) == nombre_normalizado:
            return col
    
    return None

# Mapeo de nombres de columnas antiguas a nuevas para el archivo Coste.xlsx
# Formato: nombre_nuevo: [nombres_posibles_antiguos]
MAPEO_COLUMNAS_COSTE = {
    'Artículo': ['Codigo', 'codigo', 'CODIGO', 'Código', 'codigó', 'CODIGÓ', 'Articulo', 'articulo', 'ARTICULO', 'Artículo', 'artículo', 'ARTÍCULO'],
    'Talla': ['Talla', 'talla', 'TALLA'],
    'Color': ['Color', 'color', 'COLOR'],
    'Coste': ['Coste', 'coste', 'COSTE', 'Costo', 'costo', 'COSTO', 'Precio coste', 'precio coste', 'PRECIO COSTE', 'PrecioCoste', 'preciocoste'],
    'Últ. Compra': ['Fecha ultcom', 'fecha ultcom', 'FECHA ULTCOM', 'Fecha Ult Compra', 'fecha ult compra', 'FECHA ULT COMPRA', 'Fecha última compra', 'fecha última compra', 'Fecha ultima compra', 'fecha ultima compra', 'FechaUltCompra', 'fechaultcompra', 'FECHAULTCOMPRA', 'Ult. Compra', 'ult. compra', 'ULT. COMPRA', 'Ult Compra', 'ult compra', 'ULT COMPRA']
}

def renombrar_columnas_flexible(df, mapeo):
    """
    Renombra las columnas de un DataFrame usando un mapeo flexible.
    
    Args:
        df: DataFrame a modificar
        mapeo: Diccionario con nombres nuevos como clave y lista de nombres posibles como valor
    
    Returns:
        DataFrame: DataFrame con columnas renombradas
    """
    df = df.copy()
    renombrados = {}
    
    for nombre_nuevo, nombres_posibles in mapeo.items():
        # Primero buscar coincidencia exacta (case-insensitive y sin acentos)
        for col in df.columns:
            if normalizar_texto(col) == normalizar_texto(nombre_nuevo):
                renombrados[col] = nombre_nuevo
                break
        else:
            # Si no se encuentra, buscar entre los nombres posibles
            for nombre_buscar in nombres_posibles:
                col_encontrada = buscar_columna_normalizada(df, nombre_buscar)
                if col_encontrada:
                    renombrados[col_encontrada] = nombre_nuevo
                    break
    
    # Aplicar renombrados
    if renombrados:
        df = df.rename(columns=renombrados)
    
    return df, renombrados

# ============================================================================
# CARGA DE CONFIGURACIÓN DESDE EL SISTEMA CENTRALIZADO
# ============================================================================

# Cargar configuraciones desde el sistema centralizado
ABC_CONFIG = get_abc_config()
SECCIONES = get_secciones_config()
ENCARGADOS = get_encargados_config()['ENCARGADOS']
SMTP_CONFIG = get_smtp_config()
ROTACIONES_FAMILIA = get_rotaciones_familia()
IVA_FAMILIA = get_iva_familia()
IVA_SUBFAMILIA = get_iva_subfamilia()

# Extraer configuraciones específicas (sin fechas, se calculan automáticamente)
CODIGOS_MASCOTAS_VIVO = ABC_CONFIG.get('codigos_mascotas_vivo', [])

# Configuración de formatos Excel
COLORES_RIESGO = ABC_CONFIG['colores_riesgo']
COLOR_CABECERA = ABC_CONFIG['color_cabecera']
COLOR_TEXTO_CABECERA = ABC_CONFIG['color_texto_cabecera']

# ============================================================================
# FUNCIÓN PARA ENVIAR EMAIL CON ARCHIVO ADJUNTO
# ============================================================================

def enviar_email_clasificacion(seccion: str, archivo: str, periodo: str) -> bool:
    """
    Envía un email con el archivo de clasificación ABC+D adjunto al encargado de la sección.
    
    Args:
        seccion: Nombre de la sección procesada
        archivo: Ruta completa del archivo Excel generado
        periodo: Período del análisis (formato: "dd/mm/yyyy - dd/mm/yyyy")
    
    Returns:
        bool: True si el email fue enviado exitosamente, False en caso contrario
    """
    # Obtener información del encargado
    encargado = ENCARGADOS.get(seccion.lower())
    
    if not encargado:
        print(f"  AVISO: No hay encargado configurado para la sección '{seccion}'. No se enviará email.")
        return False
    
    nombre_encargado = encargado['nombre']
    email_destinatario = encargado['email']
    
    # Verificar que el archivo existe
    if not Path(archivo).exists():
        print(f"  AVISO: El archivo '{archivo}' no existe. No se enviará email.")
        return False
    
    # Verificar contraseña en variable de entorno
    password = os.environ.get('EMAIL_PASSWORD')
    if not password:
        print(f"  AVISO: Variable de entorno 'EMAIL_PASSWORD' no configurada. No se enviará email a {nombre_encargado}.")
        return False
    
    try:
        # Crear mensaje MIME
        msg = MIMEMultipart()
        msg['From'] = f"{SMTP_CONFIG['remitente_nombre']} <{SMTP_CONFIG['remitente_email']}>"
        msg['To'] = email_destinatario
        msg['Subject'] = f"VIVEVERDE: listado ClasificacionABC+D de {seccion} del periodo {periodo}"
        
        # Cuerpo del email
        cuerpo = f"""Buenos días {nombre_encargado},

Te adjunto en este correo el listado Clasificación ABC+D de {seccion} para que lo analices y te aprendas cuales son los artículos de cada categoría:

- Artículos que no te deben faltar nunca (Categoria A).
- Artículos que confeccionan el complemento de gama (Categoría B).
- Artículos que tienen una presencia mínima en las ventas de tu sección (Categoría C).
- Artículos que no debemos tener en tienda (Categoria D).

Pon en práctica el listado.

Atentamente,

Sistema de Pedidos automáticos VIVEVERDE."""
        
        msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
        
        # Adjuntar archivo Excel
        filename = Path(archivo).name
        with open(archivo, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
        
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename= "{filename}"')
        part.add_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        msg.attach(part)
        
        # Enviar email mediante SSL
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(SMTP_CONFIG['servidor'], SMTP_CONFIG['puerto'], context=context) as server:
            server.login(SMTP_CONFIG['remitente_email'], password)
            server.sendmail(SMTP_CONFIG['remitente_email'], email_destinatario, msg.as_string())
        
        print(f"  Email enviado a {nombre_encargado} ({email_destinatario})")
        return True
        
    except smtplib.SMTPException as e:
        print(f"  ERROR SMTP al enviar email a {nombre_encargado}: {e}")
        return False
    except Exception as e:
        print(f"  ERROR al enviar email a {nombre_encargado}: {e}")
        return False

# ============================================================================
# FUNCIÓN PARA OBTENER IVA DE UN ARTÍCULO
# ============================================================================

def obtener_iva_articulo(codigo_articulo):
    """
    Obtiene el IVA correspondiente a un artículo según su familia o subfamilia.
    
    Args:
        codigo_articulo: Código del artículo (puede ser string o número)
    
    Returns:
        float: Porcentaje de IVA (10 o 21), o 21 por defecto si no se encuentra
    """
    if codigo_articulo is None or str(codigo_articulo) == 'nan':
        return 21  # IVA por defecto
    
    codigo_str = str(codigo_articulo).strip()
    
    # Eliminar decimales si viene como float
    if codigo_str.endswith('.0'):
        codigo_str = codigo_str[:-2]
    
    # Si empieza por 2, buscar en subfamilia (4 dígitos)
    if codigo_str.startswith('2'):
        if len(codigo_str) >= 4:
            subfamilia = codigo_str[:4]
            if subfamilia in IVA_SUBFAMILIA:
                return IVA_SUBFAMILIA[subfamilia]
    else:
        # Buscar en familia (2 dígitos)
        familia = codigo_str[:2]
        if familia in IVA_FAMILIA:
            return IVA_FAMILIA[familia]
    
    # IVA por defecto si no se encuentra
    return 21

# ============================================================================
# FUNCIÓN PARA DETERMINAR LA SECCIÓN DE UN ARTÍCULO
# ============================================================================

def determinar_seccion(codigo_articulo):
    """
    Determina la sección de un artículo según su código.
    
    Args:
        codigo_articulo: Código del artículo (puede ser string o número)
    
    Returns:
        str: Nombre de la sección o None si no se puede clasificar
    """
    if codigo_articulo is None:
        return None
    
    codigo_str = str(codigo_articulo).strip()
    
    # Eliminar decimales si viene como float
    if codigo_str.endswith('.0'):
        codigo_str = codigo_str[:-2]
    
    if not codigo_str or codigo_str == 'nan':
        return None
    
    # REGLA CRÍTICA: Filtrar artículos con menos de 10 dígitos
    # Esta regla tiene prioridad sobre todas las demás
    if len(codigo_str) < 10:
        return None
    
    # 1. Verificar códigos de mascotas vivos (primero, tienen prioridad)
    # Los códigos de mascotas vivos son códigos de 4 dígitos (2104, 2204, etc.)
    if codigo_str.startswith('2') and codigo_str[:4] in CODIGOS_MASCOTAS_VIVO:
        return 'mascotas_vivo'
    
    # 2. Sección 2: Mascotas manufacturadas (empieza por 2 y no está en vivos)
    if codigo_str.startswith('2'):
        return 'mascotas_manufacturado'
    
    # 3. Sección 3: Tierra/Áridos (31 o 32)
    if codigo_str.startswith('31') or codigo_str.startswith('32'):
        return 'tierra_aridos'
    
    # 4. Sección 3: Fitosanitarios (33-39)
    if codigo_str.startswith('3'):
        if len(codigo_str) >= 2:
            try:
                segundo_digito = int(codigo_str[1])
                if 3 <= segundo_digito <= 9:
                    return 'fitos'
            except (ValueError, IndexError):
                pass
    
    # 5. Secciones por primer dígito
    if codigo_str.startswith('1'):
        return 'interior'
    elif codigo_str.startswith('4'):
        return 'utiles_jardin'
    elif codigo_str.startswith('5'):
        return 'semillas'
    elif codigo_str.startswith('6'):
        return 'deco_interior'
    elif codigo_str.startswith('7'):
        return 'maf'
    elif codigo_str.startswith('8'):
        return 'vivero'
    elif codigo_str.startswith('9'):
        return 'deco_exterior'
    
    return None

# ============================================================================
# FUNCIÓN PARA PROCESAR UNA SECCIÓN ESPECÍFICA
# ============================================================================

def procesar_seccion(compras_df, ventas_df, stock_df, coste_df, nombre_seccion, seccion_info, 
                     FECHA_INICIO, FECHA_FIN, DIAS_PERIODO):
    """
    Procesa los datos de una sección específica y genera su archivo Excel.
    
    Args:
        compras_df: DataFrame de compras
        ventas_df: DataFrame de ventas
        stock_df: DataFrame de stock
        coste_df: DataFrame de costes
        nombre_seccion: Nombre de la sección a procesar
        seccion_info: Información de la sección (diccionario con descripción)
        FECHA_INICIO: Fecha de inicio del período (calculada automáticamente)
        FECHA_FIN: Fecha de fin del período (calculada automáticamente)
        DIAS_PERIODO: Número de días del período
    
    Returns:
        dict: Estadísticas del procesamiento o None si no hay datos
    """
    print(f"\n{'='*80}")
    print(f"PROCESANDO SECCIÓN: {nombre_seccion.upper()}")
    print(f"Descripción: {seccion_info['descripcion']}")
    print(f"{'='*80}")
    
    # Filtrar datos por sección
    def filtrar_por_seccion(df, columna_codigo='codigo_str'):
        """Filtra un DataFrame para incluir solo artículos de la sección"""
        if columna_codigo not in df.columns:
            return df[df[columna_codigo].apply(lambda x: determinar_seccion(x) == nombre_seccion)]
        
        # Optimizado: aplicar determinar_seccion solo a códigos únicos
        codigos_unicos = df[columna_codigo].unique()
        codigos_seccion = set()
        for codigo in codigos_unicos:
            if determinar_seccion(codigo) == nombre_seccion:
                codigos_seccion.add(codigo)
        
        return df[df[columna_codigo].isin(codigos_seccion)]
    
    # Crear copias filtradas
    compras_seccion = filtrar_por_seccion(compras_df.copy(), 'codigo_str')
    ventas_seccion = filtrar_por_seccion(ventas_df.copy(), 'codigo_str')
    stock_seccion = filtrar_por_seccion(stock_df.copy(), 'codigo_str')
    
    print(f"Datos filtrados:")
    print(f"  - Compras: {len(compras_seccion)} registros")
    print(f"  - Ventas: {len(ventas_seccion)} registros")
    print(f"  - Stock: {len(stock_seccion)} registros")
    
    # Si no hay datos en ninguna tabla, avisar y continuar
    if len(compras_seccion) == 0 and len(ventas_seccion) == 0 and len(stock_seccion) == 0:
        print(f"  AVISO: No hay datos para la sección '{nombre_seccion}'. Saltando...")
        return None
    
    # =========================================================================
    # IDENTIFICACIÓN DE ARTÍCULOS ÚNICOS
    # =========================================================================
    
    def crear_clave(row):
        return (row['codigo_str'], row['nombre_str'], row['talla_str'], row['color_str'])
    
    articulos_compras = set(compras_seccion.apply(crear_clave, axis=1))
    articulos_ventas = set(ventas_seccion.apply(crear_clave, axis=1))
    articulos_stock = set(stock_seccion.apply(crear_clave, axis=1))
    
    articulos_unicos = articulos_compras.union(articulos_ventas).union(articulos_stock)
    print(f"\nTotal artículos únicos en sección: {len(articulos_unicos)}")
    
    if len(articulos_unicos) == 0:
        print(f"  AVISO: No hay artículos únicos en la sección '{nombre_seccion}'. Saltando...")
        return None
    
    # =========================================================================
    # PROCESAMIENTO DE DATOS POR ARTÍCULO
    # =========================================================================
    
    resultados = []
    
    for clave in articulos_unicos:
        codigo, nombre, talla, color = clave
        
        # Extraer familia del código
        codigo_str = str(codigo)
        if codigo_str.startswith('2'):
            familia_codigo = codigo_str[:4]  # 4 dígitos para familias de animales
        else:
            familia_codigo = codigo_str[:2]  # 2 dígitos para el resto
        
        if familia_codigo in ROTACIONES_FAMILIA:
            nombre_familia, rotacion_familia = ROTACIONES_FAMILIA[familia_codigo]
        else:
            nombre_familia = 'OTROS'
            rotacion_familia = 90
        
        # Datos de COMPRAS
        mask_compra = (compras_seccion['codigo_str'] == codigo) & \
                      (compras_seccion['nombre_str'] == nombre) & \
                      (compras_seccion['talla_str'] == talla) & \
                      (compras_seccion['color_str'] == color)
        compras_articulo = compras_seccion[mask_compra]
        total_compras = compras_articulo['Unidades'].sum() if len(compras_articulo) > 0 else 0
        
        # Datos de VENTAS
        mask_venta = (ventas_seccion['codigo_str'] == codigo) & \
                     (ventas_seccion['nombre_str'] == nombre) & \
                     (ventas_seccion['talla_str'] == talla) & \
                     (ventas_seccion['color_str'] == color)
        ventas_articulo = ventas_seccion[mask_venta]
        unidades_vendidas = ventas_articulo['Unidades'].sum() if len(ventas_articulo) > 0 else 0
        importe_ventas = ventas_articulo['Importe'].sum() if len(ventas_articulo) > 0 else 0
        beneficio = ventas_articulo['Beneficio'].sum() if len(ventas_articulo) > 0 else 0
        coste_ventas = ventas_articulo['Coste'].sum() if len(ventas_articulo) > 0 else 0
        
        # Fecha última venta
        if len(ventas_articulo) > 0:
            ultima_venta = ventas_articulo['Fecha'].max()
            antiguedad_ultima_venta = (FECHA_FIN - ultima_venta).days
        else:
            antiguedad_ultima_venta = DIAS_PERIODO
        
        # Datos de STOCK
        mask_stock = (stock_seccion['codigo_str'] == codigo) & \
                     (stock_seccion['nombre_str'] == nombre) & \
                     (stock_seccion['talla_str'] == talla) & \
                     (stock_seccion['color_str'] == color)
        stock_articulo = stock_seccion[mask_stock]
        stock_inicial = stock_articulo['Unidades'].sum() if len(stock_articulo) > 0 else 0
        precio_coste_stock = stock_articulo['Precio'].iloc[0] if len(stock_articulo) > 0 else 0
        
        # Métricas
        stock_disponible_total = stock_inicial + total_compras
        stock_final = stock_inicial + total_compras - unidades_vendidas
        
        # Tasa de Venta
        if stock_disponible_total > 0:
            tasa_venta = (unidades_vendidas / stock_disponible_total) * 100
        else:
            tasa_venta = 0
        
        # Antigüedad Stock
        if stock_final > 0:
            if stock_inicial - unidades_vendidas > 0:
                antiguedad_stock = DIAS_PERIODO
                origen_stock = 'Stock inicial'
            else:
                ventas_acumuladas = 0
                compras_ordenadas = compras_articulo.sort_values('Fecha')
                for idx, compra in compras_ordenadas.iterrows():
                    ventas_acumuladas += compra['Unidades']
                    if ventas_acumuladas >= (stock_inicial + total_compras - stock_final):
                        antiguedad_stock = (FECHA_FIN - compra['Fecha']).days
                        origen_stock = f'Compra {compra["Fecha"].strftime("%d/%m/%Y")}'
                        break
                else:
                    if len(compras_ordenadas) > 0:
                        ultima_compra = compras_ordenadas.iloc[-1]
                        antiguedad_stock = (FECHA_FIN - ultima_compra['Fecha']).days
                        origen_stock = f'Compra {ultima_compra["Fecha"].strftime("%d/%m/%Y")}'
                    else:
                        antiguedad_stock = DIAS_PERIODO
                        origen_stock = 'Stock inicial'
        else:
            antiguedad_stock = 0
            origen_stock = 'Sin stock'
        
        # % Rotación Consumida
        if stock_final > 0 and rotacion_familia > 0:
            pct_rotacion_consumida = (antiguedad_stock / rotacion_familia) * 100
        else:
            pct_rotacion_consumida = 0
        
        # Descuento Sugerido
        if pct_rotacion_consumida <= 65:
            descuento_sugerido = 0
        elif pct_rotacion_consumida <= 100:
            descuento_sugerido = 10
        elif pct_rotacion_consumida <= 150:
            descuento_sugerido = 20
        else:
            descuento_sugerido = 30
        
        # Riesgo de Merma/Inmovilizado
        es_categoria_d = (unidades_vendidas == 0)
        
        if es_categoria_d:
            if stock_final == 0:
                riesgo = 'Cero'
            else:
                riesgo = 'Crítico'
        else:
            if stock_final == 0:
                riesgo = 'Cero'
            elif pct_rotacion_consumida <= 65:
                riesgo = 'Bajo'
            elif pct_rotacion_consumida <= 100:
                riesgo = 'Medio'
            elif pct_rotacion_consumida <= 150:
                riesgo = 'Alto'
            else:
                riesgo = 'Crítico'
        
        # Rotación Excedida
        if antiguedad_ultima_venta > rotacion_familia and stock_final > 0:
            rotacion_excedida = stock_final
        else:
            rotacion_excedida = 0
        
        # Clasificación por Stock Final
        demanda_mensual_promedio = unidades_vendidas / 2
        if stock_final == 0:
            nivel_stock = 'Cero'
        elif stock_final <= demanda_mensual_promedio * 0.5:
            nivel_stock = 'Bajo'
        elif stock_final <= demanda_mensual_promedio:
            nivel_stock = 'Normal'
        else:
            nivel_stock = 'Elevado'
        
        # Ventas media diaria
        ventas_media_diaria = unidades_vendidas / DIAS_PERIODO if DIAS_PERIODO > 0 else 0
        
        # Stock Mínimo
        if rotacion_familia == 7:
            stock_minimo = ventas_media_diaria * 3.5
        elif rotacion_familia == 15:
            stock_minimo = ventas_media_diaria * 7.5
        elif rotacion_familia == 30:
            stock_minimo = ventas_media_diaria * 15
        elif rotacion_familia == 60:
            stock_minimo = ventas_media_diaria * 30
        elif rotacion_familia == 90:
            stock_minimo = ventas_media_diaria * 45
        else:
            stock_minimo = ventas_media_diaria * 45
        
        # Stock Máximo
        if rotacion_familia == 7:
            stock_maximo = ventas_media_diaria * 10.5
        elif rotacion_familia == 15:
            stock_maximo = ventas_media_diaria * 22.5
        elif rotacion_familia == 30:
            stock_maximo = ventas_media_diaria * 45
        elif rotacion_familia == 60:
            stock_maximo = ventas_media_diaria * 90
        elif rotacion_familia == 90:
            stock_maximo = ventas_media_diaria * 135
        else:
            stock_maximo = ventas_media_diaria * 135
        
        # Días de cobertura
        if ventas_media_diaria > 0:
            dias_cobertura = stock_final / ventas_media_diaria
        else:
            dias_cobertura = 0
        
        resultados.append({
            'Artículo': codigo,
            'Nombre artículo': nombre,
            'Talla': talla if talla else '',
            'Color': color if color else '',
            'Familia': familia_codigo,
            'Nombre Familia': nombre_familia,
            'Rotación Familia (días)': rotacion_familia,
            'Stock Inicial (unidades)': stock_inicial,
            'Compras Período (unidades)': total_compras,
            'Ventas (unidades)': unidades_vendidas,
            'Importe ventas (€)': round(importe_ventas, 2),
            'Beneficio (importe €)': round(beneficio, 2),
            'Coste Ventas Real (€)': round(coste_ventas, 2),
            'Stock Disponible Total': stock_disponible_total,
            'Tasa de venta (%)': round(tasa_venta, 2),
            'Rotación excedida (unidades)': rotacion_excedida,
            'Stock mínimo (unidades)': round(stock_minimo, 1),
            'Stock máximo (unidades)': round(stock_maximo, 1),
            'Stock Final (unidades)': stock_final,
            'Antigüedad Última Venta (días)': antiguedad_ultima_venta,
            'Antigüedad Stock (días)': antiguedad_stock,
            '% Rotación Consumido': round(pct_rotacion_consumida, 2),
            'Descuento Sugerido (%)': descuento_sugerido,
            'Riesgo de Merma/ inmovilizado': riesgo,
            'Nivel Stock Final': nivel_stock,
            'Días de cobertura': round(dias_cobertura, 1),
            'Origen Stock Final': origen_stock,
            'Precio Coste Unitario (€)': precio_coste_stock,
        })
    
    df_resultados = pd.DataFrame(resultados)
    print(f"\nTotal artículos procesados: {len(df_resultados)}")
    
    # =========================================================================
    # CLASIFICACIÓN ABC+D
    # =========================================================================
    
    df_con_ventas = df_resultados[df_resultados['Coste Ventas Real (€)'] > 0].copy()
    df_sin_ventas = df_resultados[df_resultados['Coste Ventas Real (€)'] == 0].copy()
    
    print(f"Artículos con ventas: {len(df_con_ventas)}")
    print(f"Artículos sin ventas: {len(df_sin_ventas)}")
    
    if len(df_con_ventas) > 0:
        df_con_ventas = df_con_ventas.sort_values('Coste Ventas Real (€)', ascending=False)
        
        total_coste = df_con_ventas['Coste Ventas Real (€)'].sum()
        df_con_ventas['% Individual'] = (df_con_ventas['Coste Ventas Real (€)'] / total_coste) * 100
        df_con_ventas['% Acumulado'] = df_con_ventas['% Individual'].cumsum()
        
        def asignar_categoria(pct_acumulado):
            if pct_acumulado <= 80:
                return 'A'
            elif pct_acumulado <= 95:
                return 'B'
            else:
                return 'C'
        
        df_con_ventas['Categoria ABC'] = df_con_ventas['% Acumulado'].apply(asignar_categoria)
        
        print(f"\n  Categoría A: {len(df_con_ventas[df_con_ventas['Categoria ABC'] == 'A'])} artículos")
        print(f"  Categoría B: {len(df_con_ventas[df_con_ventas['Categoria ABC'] == 'B'])} artículos")
        print(f"  Categoría C: {len(df_con_ventas[df_con_ventas['Categoria ABC'] == 'C'])} artículos")
    
    df_sin_ventas['Categoria ABC'] = 'D'
    
    df_clasificado = pd.concat([df_con_ventas, df_sin_ventas], ignore_index=True)
    
    print(f"\n  Categoría D: {len(df_clasificado[df_clasificado['Categoria ABC'] == 'D'])} artículos")
    
    # =========================================================================
    # ASIGNACIÓN DE ESCENARIOS
    # =========================================================================
    
    def asignar_escenario(row):
        stock_final = row['Stock Final (unidades)']
        pct_rotacion = row['% Rotación Consumido']
        antiguedad_venta = row['Antigüedad Última Venta (días)']
        rotacion = row['Rotación Familia (días)']
        categoria = row['Categoria ABC']
        riesgo = row['Riesgo de Merma/ inmovilizado']
        nivel_stock = row['Nivel Stock Final']
        
        if stock_final == 0:
            if rotacion > 0:
                pct_rotacion_venta = (antiguedad_venta / rotacion) * 100
            else:
                pct_rotacion_venta = 0
            
            if categoria in ['A', 'B']:
                if pct_rotacion_venta <= 24: return '13A'
                elif pct_rotacion_venta <= 50: return '13B'
                elif pct_rotacion_venta <= 100: return '13C'
                else: return '13D'
            else:
                if pct_rotacion_venta <= 24: return '26A'
                elif pct_rotacion_venta <= 50: return '26B'
                elif pct_rotacion_venta <= 100: return '26C'
                else: return '26D'
        else:
            if nivel_stock == 'Elevado':
                if riesgo == 'Crítico': return '1' if categoria in ['A', 'B'] else '14'
                elif riesgo == 'Alto': return '2' if categoria in ['A', 'B'] else '15'
                elif riesgo == 'Medio': return '3' if categoria in ['A', 'B'] else '16'
                else: return '4' if categoria in ['A', 'B'] else '17'
            elif nivel_stock == 'Normal':
                if riesgo == 'Crítico': return '5' if categoria in ['A', 'B'] else '18'
                elif riesgo == 'Alto': return '6' if categoria in ['A', 'B'] else '19'
                elif riesgo == 'Medio': return '7' if categoria in ['A', 'B'] else '20'
                else: return '8' if categoria in ['A', 'B'] else '21'
            else:
                if riesgo == 'Crítico': return '9' if categoria in ['A', 'B'] else '22'
                elif riesgo == 'Alto': return '10' if categoria in ['A', 'B'] else '23'
                elif riesgo == 'Medio': return '11' if categoria in ['A', 'B'] else '24'
                else: return '12' if categoria in ['A', 'B'] else '25'
    
    df_clasificado['Escenario'] = df_clasificado.apply(asignar_escenario, axis=1)
    
    # =========================================================================
    # ACCIONES SUGERIDAS
    # =========================================================================
    
    TEXTOS_ESCENARIOS = {
        '1': "DESCUENTO MÁXIMO + REDUCCIÓN COMPRAS: Aplicar descuento [descuento]% inmediato. Reducir compras 50% próxima temporada. Stock objetivo: [unidades] unidades. Prioridad alta.",
        '2': "DESCUENTO MODERADO + REDUCCIÓN COMPRAS: Aplicar descuento [descuento]% para dinamizar ventas. Reducir compras 35% próxima temporada. Stock objetivo: [unidades] unidades. Monitorear.",
        '3': "DESCUENTO PREVENTIVO + AJUSTE COMPRAS: Aplicar descuento [descuento]% para anticipar venta. Reducir compras 20% próxima temporada. Mantener bajo observación semanal.",
        '4': "MANTENER + GESTIÓN ACTIVA: Stock fresco de calidad. Reducir compras 15% próxima temporada. Stock actual suficiente para [X días] días.",
        '5': "DESCUENTO CORRECTIVO + MONITOREO: Aplicar descuento [descuento]% a stock actual para renovar inventario. Mantener nivel de compras actual.",
        '6': "DESCUENTO LEVE + OPTIMIZACIÓN: Aplicar descuento [descuento]% para renovar inventario. Reducir compras 15% próxima temporada.",
        '7': "OPTIMIZAR PREVENTIVO: Aplicar descuento [descuento]% preventivo. Mantener nivel de compras actual. Stock bien gestionado.",
        '8': "MANTENER ESTRATEGIA ACTUAL: Gestión excelente. Stock óptimo y fresco. Mantener nivel de compras actual. Producto clave del catálogo.",
        '9': "INVESTIGAR + REDISEÑAR: Analizar causa de baja rotación. Mantener stock mínimo. Implementar acciones de venta. Reducir compras 25%.",
        '10': "PROMOCIÓN ACTIVA + AJUSTE: Implementar promoción del 15% para estimula demanda. Aumentar visibilidad en punto de venta.",
        '11': "REPOSICIÓN SELECTIVA: Aumentar compras 15% para evitar ruptura de stock. Aplicar descuento 5% para consolidar demanda.",
        '12': "AUMENTAR STOCK: Producto de alto interés. Incrementar compras 20% próxima temporada. Stock actual: [unidades] unidades. Maximizar disponibilidad.",
        '13A': "URGENTE - REPOSICIÓN INMEDIATA: Producto de alta demanda agotado. Recompra prioritaria inmediata. Aumentar compras 40%. Stock objetivo: [unidades] unidades.",
        '13B': "REPOSICIÓN PRIORITARIA: Producto agotado con demanda reciente. Aumentar compras 25%. Stock objetivo: [unidades] unidades.",
        '13C': "REPOSICIÓN PROGRAMADA: Stock agotado con rotación moderada. Mantener nivel de compras anterior. Stock objetivo: [unidades] unidades.",
        '13D': "EVALUAR CONTINUIDAD: Producto agotado con demanda decreciente. Reducir compras 30% próxima temporada. Evaluar continuidad en catálogo.",
        '14': "LIQUIDACIÓN URGENTE: Aplicar descuento [descuento]% inmediato. Eliminar del catálogo próxima temporada. Capital liberado: [importe]€. Prioridad máxima.",
        '15': "REDUCCIÓN AGRESIVA: Aplicar descuento [descuento]% inmediato. Reducir compras 70% próxima temporada. Stock objetivo: [unidades] unidades. Riesgo alto de merma.",
        '16': "DESCUENTO PREVENTIVO: Aplicar descuento [descuento]% para acelerar rotación. Reducir compras 40% próxima temporada. Monitorear evolución semanal.",
        '17': "MANTENER SIN DESCUENTO: Stock fresco de calidad. Reducir compras 25% próxima temporada. Stock actual suficiente para [X días] días.",
        '18': "LIQUIDACIÓN PARCIAL: Aplicar descuento [descuento]% a stock actual. Reducir compras 50% próxima temporada. Producto de baja rotación confirmada.",
        '19': "DESCUENTO MODERADO: Aplicar descuento [descuento]% para renovar inventario. Reducir compras 30% próxima temporada. Stock actual en rango aceptable pero envejecido.",
        '20': "OPTIMIZAR: Aplicar descuento [descuento]% preventivo. Mantener nivel de compras actual. Stock bien gestionado. Continuar monitoreo.",
        '21': "MANTENER ESTRATEGIA ACTUAL: Gestión excelente. Stock óptimo y fresco. Mantener nivel de compras. Producto bien equilibrado.",
        '22': "ELIMINAR DEL CATÁLOGO: Aplicar descuento [descuento]% para liquidar stock residual. NO recomprar. Bajo interés confirmado del cliente.",
        '23': "LIQUIDAR Y DESCATALOGAR: Aplicar descuento [descuento]% para agotar stock. NO recomprar próxima temporada. Producto sin demanda suficiente.",
        '24': "COMPRAS CONSERVADORAS: Aplicar descuento [descuento]% al stock actual. Reducir compras 50% próxima temporada. Demanda limitada confirmada.",
        '25': "AUMENTAR STOCK: Producto de alto interés. Incrementar compras 30% próxima temporada. Stock actual: [unidades] unidades. Alta rotación confirmada.",
        '26A': "URGENTE - RUPTURA DE STOCK: Producto de alta demanda agotado. Recompra inmediata prioritaria. Aumentar compras 50%. Stock objetivo: [unidades] unidades. Pérdida de ventas estimada.",
        '26B': "RECOMPRA PRIORITARIA: Producto agotado con demanda reciente. Aumentar compras 30%. Stock objetivo: [unidades] unidades. Monitorear demanda próximas semanas.",
        '26C': "RECOMPRA MODERADA: Stock agotado con rotación moderada. Mantener nivel de compras anterior. Stock objetivo: [unidades] unidades. Demanda estable.",
        '26D': "RECOMPRA CONSERVADORA: Producto agotado de baja rotación. Reducir compras 40% próxima temporada. Stock objetivo mínimo: [unidades] unidades.",
    }
    
    def generar_accion_sugerida(row):
        escenario = row['Escenario']
        if escenario not in TEXTOS_ESCENARIOS:
            return "Sin acción asignada"
        
        texto = TEXTOS_ESCENARIOS[escenario]
        descuento = row['Descuento Sugerido (%)']
        stock_final = row['Stock Final (unidades)']
        stock_minimo = row['Stock mínimo (unidades)']
        dias_cobertura = row['Días de cobertura']
        
        stock_objetivo_14_dias = max(1, round(stock_minimo * 0.5, 0))
        stock_objetivo_aumentar = max(1, round(stock_final * 1.5, 0))
        stock_objetivo_doble = max(1, round(stock_final * 2, 0))
        capital_liberado = round(stock_final * row['Precio Coste Unitario (€)'] * 0.7, 2)
        
        texto = texto.replace('[descuento]', str(descuento))
        texto = texto.replace('[unidades]', str(int(stock_objetivo_14_dias)))
        texto = texto.replace('[X días]', str(int(dias_cobertura)))
        texto = texto.replace('[importe]', str(capital_liberado))
        
        # Casos especiales
        if escenario == '12':
            texto = f"AUMENTAR STOCK: Producto de alto interés. Incrementar compras 20% próxima temporada. Stock actual: {int(stock_final)} unidades. Stock objetivo: {int(stock_objetivo_aumentar)} unidades. Maximizar disponibilidad."
        elif escenario == '13A':
            texto = f"URGENTE - REPOSICIÓN INMEDIATA: Producto de alta demanda agotado. Recompra prioritaria inmediata. Aumentar compras 40%. Stock objetivo: {int(stock_objetivo_doble)} unidades. Evitar futura ruptura."
        elif escenario == '13B':
            texto = f"REPOSICIÓN PRIORITARIA: Producto agotado con demanda reciente. Aumentar compras 25%. Stock objetivo: {int(stock_objetivo_aumentar)} unidades. Programar reposición para próxima semana."
        elif escenario == '25':
            texto = f"AUMENTAR STOCK: Producto de alto interés. Incrementar compras 30% próxima temporada. Stock actual: {int(stock_final)} unidades. Stock objetivo: {int(stock_objetivo_aumentar)} unidades. Alta rotación confirmada."
        elif escenario == '26A':
            texto = f"URGENTE - RUPTURA DE STOCK: Producto de alta demanda agotado. Recompra inmediata prioritaria. Aumentar compras 50%. Stock objetivo: {int(stock_objetivo_doble)} unidades. Pérdida de ventas estimada."
        elif escenario == '26B':
            texto = f"RECOMPRA PRIORITARIA: Producto agotado con demanda reciente. Aumentar compras 30%. Stock objetivo: {int(stock_objetivo_aumentar)} unidades. Monitorear demanda próximas semanas."
        
        return texto
    
    df_clasificado['Acción Sugerida'] = df_clasificado.apply(generar_accion_sugerida, axis=1)
    
    # =========================================================================
    # SEPARACIÓN POR CATEGORÍAS
    # =========================================================================
    
    columnas_salida = [
        'Artículo', 'Nombre artículo', 'Talla', 'Color',
        'Familia', 'Nombre Familia', 'Rotación Familia (días)',
        'Ventas (unidades)', 'Importe ventas (€)', 'Beneficio (importe €)',
        'Tasa de venta (%)', 'Rotación excedida (unidades)',
        'Stock mínimo (unidades)', 'Stock máximo (unidades)',
        'Stock Final (unidades)', 'Antigüedad Última Venta (días)',
        'Antigüedad Stock (días)', '% Rotación Consumido',
        'Descuento Sugerido (%)', 'Riesgo de Merma/ inmovilizado',
        'Acción Sugerida', 'Origen Stock Final', 'Escenario'
    ]
    
    df_categoria_a = df_clasificado[df_clasificado['Categoria ABC'] == 'A'][columnas_salida].copy()
    df_categoria_b = df_clasificado[df_clasificado['Categoria ABC'] == 'B'][columnas_salida].copy()
    df_categoria_c = df_clasificado[df_clasificado['Categoria ABC'] == 'C'][columnas_salida].copy()
    df_categoria_d = df_clasificado[df_clasificado['Categoria ABC'] == 'D'][columnas_salida].copy()
    
    # =========================================================================
    # GUARDAR ARCHIVO EXCEL
    # =========================================================================
    
    # Generar nombre de archivo con período
    periodo_str = f"{FECHA_INICIO.strftime('%Y%m%d')}-{FECHA_FIN.strftime('%Y%m%d')}"
    nombre_archivo = f"data/input/CLASIFICACION_ABC+D_{nombre_seccion.upper()}_{periodo_str}.xlsx"
    
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df_categoria_a.to_excel(writer, sheet_name='CATEGORIA A – BASICOS', index=False)
        df_categoria_b.to_excel(writer, sheet_name='CATEGORIA B – COMPLEMENTO', index=False)
        df_categoria_c.to_excel(writer, sheet_name='CATEGORIA C – BAJO IMPACTO', index=False)
        df_categoria_d.to_excel(writer, sheet_name='CATEGORIA D – SIN VENTAS', index=False)
    
    print(f"\nArchivo generado: {nombre_archivo}")
    
    # =========================================================================
    # APLICAR FORMATOS EXCEL
    # =========================================================================
    
    def aplicar_formato_hoja(worksheet, df):
        worksheet.column_dimensions['A'].width = 18
        worksheet.column_dimensions['B'].width = 45
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['M'].width = 18
        worksheet.column_dimensions['N'].width = 18
        worksheet.column_dimensions['S'].width = 22
        worksheet.column_dimensions['U'].width = 32
        worksheet.column_dimensions['X'].width = 15
        
        columnas_ocultar = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'O', 'P', 'Q', 'R', 'V', 'W']
        for col in columnas_ocultar:
            worksheet.column_dimensions[col].hidden = True
        
        fill_cabecera = PatternFill(start_color=COLOR_CABECERA, end_color=COLOR_CABECERA, fill_type='solid')
        font_cabecera = Font(color=COLOR_TEXTO_CABECERA, bold=True, size=10)
        
        for cell in worksheet[1]:
            cell.fill = fill_cabecera
            cell.font = font_cabecera
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        worksheet.row_dimensions[1].height = 45
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        columnas_centradas = ['M', 'N', 'S', 'T']
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = thin_border
                
                col_letter = chr(64 + col_idx)
                if col_letter in columnas_centradas:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_letter in ['A', 'B', 'C', 'D', 'U']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                if col_letter == 'T' and cell.value:
                    riesgo = str(cell.value)
                    if riesgo in COLORES_RIESGO:
                        cell.fill = PatternFill(start_color=COLORES_RIESGO[riesgo], 
                                               end_color=COLORES_RIESGO[riesgo], 
                                               fill_type='solid')
        
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.margin_left = 0
        worksheet.page_setup.margin_right = 0
        worksheet.page_setup.margin_top = 0
        worksheet.page_setup.margin_bottom = 0
        
        worksheet.auto_filter.ref = worksheet.dimensions
        
        return worksheet
    
    wb = load_workbook(nombre_archivo)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df = pd.read_excel(nombre_archivo, sheet_name=sheet_name)
        aplicar_formato_hoja(ws, df)
    
    wb.save(nombre_archivo)
    
    print(f"\nEnviando email al encargado de la sección...")
    
    # Formatear período para el email
    periodo_str = f"{FECHA_INICIO.strftime('%d/%m/%Y')} - {FECHA_FIN.strftime('%d/%m/%Y')}"
    
    # Enviar email con el archivo adjunto
    email_enviado = enviar_email_clasificacion(nombre_seccion, nombre_archivo, periodo_str)
    
    if email_enviado:
        print(f"  ✓ Email enviado correctamente a {ENCARGADOS.get(nombre_seccion.lower(), {}).get('nombre', nombre_seccion)}")
    else:
        print(f"  ✗ No se pudo enviar el email a {ENCARGADOS.get(nombre_seccion.lower(), {}).get('nombre', nombre_seccion)}")
    
    # Retornar estadísticas
    return {
        'seccion': nombre_seccion,
        'archivo': nombre_archivo,
        'total_articulos': len(df_clasificado),
        'categoria_a': len(df_categoria_a),
        'categoria_b': len(df_categoria_b),
        'categoria_c': len(df_categoria_c),
        'categoria_d': len(df_categoria_d),
        'email_enviado': email_enviado,
    }

# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def main():
    """Función principal del script"""
    
    # Parsear argumentos de línea de comandos
    parser = argparse.ArgumentParser(
        description='Motor de Cálculo ABC+D para Gestión de Inventarios - Versión V2',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Períodos disponibles:
  --periodo 1: Enero - Febrero (59 días)
  --periodo 2: Marzo - Mayo (92 días)
  --periodo 3: Junio - Agosto (92 días)
  --periodo 4: Septiembre - Diciembre (122 días)

Ejemplos de uso:
  python clasificacionABC.py --periodo 1              # Procesa Período 1 completo
  python clasificacionABC.py --periodo 2 --seccion vivero  # Solo sección vivero del período 2
  python clasificacionABC.py --periodo 3 --verbose         # Modo verbose para período 3

Secciones disponibles:
  interior, utiles_jardin, semillas, deco_interior, maf, vivero, 
  deco_exterior, mascotas_manufacturado, mascotas_vivo, tierra_aridos, fitos
        """
    )
    parser.add_argument(
        '-p', '--periodo',
        type=int,
        help='Número del período a procesar (1-4). Si no se especifica, usa fechas de Ventas.xlsx'
    )
    parser.add_argument(
        '-s', '--seccion',
        type=str,
        help='Procesar solo una sección específica (opcional)'
    )
    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Activar modo verbose con logs detallados'
    )
    args = parser.parse_args()
    
    # Validar período si se especificó
    if args.periodo and args.periodo not in DEFINICION_PERIODOS:
        print(f"ERROR: Período '{args.periodo}' no válido.")
        print(f"Períodos disponibles: {', '.join(map(str, DEFINICION_PERIODOS.keys()))}")
        sys.exit(1)
    
    seccion_especifica = args.seccion.lower() if args.seccion else None
    
    # Validar sección si se especificó
    if seccion_especifica and seccion_especifica not in SECCIONES:
        print(f"ERROR: Sección '{seccion_especifica}' no válida.")
        print(f"Secciones disponibles: {', '.join(sorted(SECCIONES.keys()))}")
        sys.exit(1)
    
    print("=" * 80)
    print("MOTOR DE CÁLCULO ABC+D PARA GESTIÓN DE INVENTARIOS - V2")
    print("=" * 80)
    
    if args.periodo:
        print(f"\nMODO: Por Período")
        print(f"Período seleccionado: {DEFINICION_PERIODOS[args.periodo]['descripcion']}")
    elif seccion_especifica:
        print(f"\nMODO: Mono-sección")
        print(f"Sección seleccionada: {seccion_especifica}")
    else:
        print(f"\nMODO: Multi-sección (todas las secciones)")
    
    # =========================================================================
    # CARGA DE DATOS DESDE ARCHIVOS SEPARADOS
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 1: CARGA Y EXTRACCIÓN DE DATOS")
    print("=" * 80)
    
    try:
        # Cargar archivos del ERP con soporte para timestamps y case-insensitive
        import glob
        
        def buscar_archivo_erp(nombre_base):
            """Busca archivo con múltiples variaciones de mayúsculas/minúsculas"""
            variaciones = [
                nombre_base,
                nombre_base.lower(),
                nombre_base.capitalize(),
                nombre_base.upper()
            ]
            for var in variaciones:
                # Primero intentar con timestamp
                archivo = find_latest_file('data/input', var)
                if archivo:
                    return archivo
            
            # Si no encontramos con timestamp, buscar archivos que coincidan case-insensitive
            base_lower = nombre_base.lower()
            for archivo in glob.glob('data/input/*.xlsx'):
                nombre_archivo = os.path.basename(archivo)
                if nombre_archivo.lower().startswith(base_lower.lower()):
                    return archivo
            
            return None

        archivo_compras = buscar_archivo_erp('SPA_Compras')
        if archivo_compras:
            compras_df = pd.read_excel(archivo_compras)
            print(f"COMPRAS: {os.path.basename(archivo_compras)}")
        else:
            compras_df = pd.read_excel('data/input/compras.xlsx')
            print("COMPRAS: compras.xlsx (formato legacy)")

        archivo_ventas = buscar_archivo_erp('SPA_Ventas')
        if archivo_ventas:
            ventas_df = pd.read_excel(archivo_ventas)
            print(f"VENTAS: {os.path.basename(archivo_ventas)}")
        else:
            ventas_df = pd.read_excel('data/input/Ventas.xlsx')
            print("VENTAS: Ventas.xlsx (formato legacy)")

        archivo_stock = buscar_archivo_erp('SPA_Stock')
        if archivo_stock:
            stock_df = pd.read_excel(archivo_stock)
            print(f"STOCK: {os.path.basename(archivo_stock)}")
        else:
            stock_df = pd.read_excel('data/input/Stock.xlsx')
            print("STOCK: Stock.xlsx (formato legacy)")

        archivo_coste = buscar_archivo_erp('SPA_Coste')
        if archivo_coste:
            coste_df = pd.read_excel(archivo_coste)
            print(f"COSTE: {os.path.basename(archivo_coste)}")
        else:
            coste_df = pd.read_excel('data/input/Coste.xlsx')
            print("COSTE: Coste.xlsx (formato legacy)")
        
        # Aplicar renombrado flexible de columnas para el archivo Coste.xlsx
        coste_df, columnas_renombradas = renombrar_columnas_flexible(coste_df, MAPEO_COLUMNAS_COSTE)
        if columnas_renombradas:
            print(f"COSTE: Columnas renombradas automáticamente: {columnas_renombradas}")
        
        # Verificar que las columnas esenciales existan después del renombrado
        columnas_esenciales = ['Artículo', 'Talla', 'Color', 'Coste']
        columnas_faltantes = [col for col in columnas_esenciales if col not in coste_df.columns]
        if columnas_faltantes:
            print(f"ERROR: Faltan columnas esenciales en Coste.xlsx después del renombrado: {columnas_faltantes}")
            print(f"Columnas disponibles en Coste.xlsx: {list(coste_df.columns)}")
            sys.exit(1)
            
    except FileNotFoundError as e:
        print(f"ERROR: No se encontró el archivo: {e.filename}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR al cargar archivos: {e}")
        sys.exit(1)
    
    print(f"COMPRAS: {len(compras_df)} registros cargados")
    print(f"VENTAS: {len(ventas_df)} registros cargados")
    print(f"STOCK: {len(stock_df)} registros cargados")
    print(f"COSTE: {len(coste_df)} registros cargados")
    
    # =========================================================================
    # CÁLCULO DEL PERÍODO DE ANÁLISIS
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 1A: DETERMINACIÓN DEL PERÍODO DE ANÁLISIS")
    print("=" * 80)
    
    if args.periodo:
        # Usar período predefinido
        periodo_info = DEFINICION_PERIODOS[args.periodo]
        año_actual = datetime.now().year
        fecha_inicio_str = f"{periodo_info['fecha_inicio']}/{año_actual}"
        fecha_fin_str = f"{periodo_info['fecha_fin']}/{año_actual}"
        FECHA_INICIO = datetime.strptime(fecha_inicio_str, '%d/%m/%Y')
        FECHA_FIN = datetime.strptime(fecha_fin_str, '%d/%m/%Y')
        DIAS_PERIODO = periodo_info['dias']
        
        print(f"\nPeríodo configurado manualmente:")
        print(f"   Período: {periodo_info['nombre']} ({periodo_info['descripcion']})")
        print(f"   Desde: {FECHA_INICIO.strftime('%d de %B de %Y')}")
        print(f"   Hasta: {FECHA_FIN.strftime('%d de %B de %Y')}")
        print(f"   Días: {DIAS_PERIODO}")
    else:
        # Calcular automáticamente las fechas mínima y máxima del archivo de ventas
        FECHA_INICIO, FECHA_FIN, DIAS_PERIODO = calcular_periodo_ventas(ventas_df)
        
        print(f"\nPeríodo calculado automáticamente desde Ventas.xlsx:")
        print(f"   Desde: {FECHA_INICIO.strftime('%d de %B de %Y')}")
        print(f"   Hasta: {FECHA_FIN.strftime('%d de %B de %Y')}")
        print(f"   Días: {DIAS_PERIODO}")
    
    # Filtrar filas con Artículo vacío en Compras
    filas_antes = len(compras_df)
    compras_df = compras_df[compras_df['Artículo'].notna() & (compras_df['Artículo'] != '')]
    filas_eliminadas = filas_antes - len(compras_df)
    if filas_eliminadas > 0:
        print(f"Eliminadas {filas_eliminadas} filas con artículo vacío en Compras")
    
    # Rellenar celdas vacías en STOCK
    filas_vacias_stock = stock_df['Artículo'].isna().sum()
    if filas_vacias_stock > 0:
        stock_df['Artículo'] = stock_df['Artículo'].ffill()
        stock_df['Nombre artículo'] = stock_df['Nombre artículo'].ffill()
        print(f"STOCK: {len(stock_df)} registros ({filas_vacias_stock} celdas vacías preenchidas)")
    else:
        print(f"STOCK: {len(stock_df)} registros")
    
    # =========================================================================
    # PROCESAR DATOS DE VENTAS - Calcular Coste y Beneficio
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 1B: CÁLCULO DE COSTE Y BENEFICIO EN VENTAS")
    print("=" * 80)
    
    # Filtrar solo filas de tipo 'Detalle'
    filas_ventas_total = len(ventas_df)
    ventas_df = ventas_df[ventas_df['Tipo registro'] == 'Detalle'].copy()
    print(f"VENTAS: {filas_ventas_total} filas totales → {len(ventas_df)} filas de Detalle")
    
    # Normalizar claves de unión en Coste
    # Usar la columna 'Últ. Compra' si existe, si no usar cualquier columna de fecha disponible
    columna_fecha = buscar_columna_normalizada(coste_df, 'Últ. Compra')
    if columna_fecha:
        coste_df_sorted = coste_df.sort_values(columna_fecha, ascending=False)
    else:
        # Buscar cualquier columna que parezca una fecha
        for col in coste_df.columns:
            if 'fecha' in normalizar_texto(col) or 'date' in normalizar_texto(col).lower():
                print(f"  AVISO: Usando columna '{col}' como fecha de última compra")
                coste_df_sorted = coste_df.sort_values(col, ascending=False)
                break
        else:
            # Si no hay columna de fecha, no ordenar
            coste_df_sorted = coste_df.copy()
    
    # Usar las nuevas columnas renombradas: Artículo, Talla, Color
    coste_df_latest = coste_df_sorted.drop_duplicates(subset=['Artículo', 'Talla', 'Color'], keep='first').copy()
    
    def normalize_keys(df):
        df = df.copy()
        df['Artículo'] = df['Artículo'].astype(str).str.replace(r'\.0$', '', regex=True)
        df['Talla'] = df['Talla'].fillna('').astype(str).str.strip()
        df['Color'] = df['Color'].fillna('').astype(str).str.strip()
        return df
    
    def normalize_keys_coste(df):
        """Normalizar claves para el archivo Coste.xlsx (usa 'Artículo' en lugar de 'Codigo')"""
        df = df.copy()
        # Usar la columna 'Artículo' que ya fue renombrada del original 'Codigo'
        df['Artículo'] = df['Artículo'].astype(str).str.replace(r'\.0$', '', regex=True)
        df['Talla'] = df['Talla'].fillna('').astype(str).str.strip()
        df['Color'] = df['Color'].fillna('').astype(str).str.strip()
        return df
    
    ventas_normalized = normalize_keys(ventas_df)
    coste_normalized = normalize_keys_coste(coste_df_latest)
    
    # Seleccionar solo las columnas necesarias de coste (ya renombrado a Artículo)
    coste_for_merge = coste_normalized[['Artículo', 'Talla', 'Color', 'Coste']].copy()
    
    # Merge de ventas con costes
    ventas_with_costs = pd.merge(
        ventas_normalized,
        coste_for_merge,
        on=['Artículo', 'Talla', 'Color'],
        how='left'
    )
    
    # Calcular Coste total
    def calculate_cost(row):
        try:
            unidades = row['Unidades'] if pd.notna(row['Unidades']) else 1
            importe = row['Importe'] if pd.notna(row['Importe']) else 0
            coste_unitario = row['Coste'] if pd.notna(row['Coste']) else 0
            
            if coste_unitario > 0:
                return unidades * coste_unitario
            
            if unidades > 0 and importe > 0:
                pvp = importe / unidades
                iva = obtener_iva_articulo(row['Artículo'])
                
                if iva == 10:
                    coste_calculado = (pvp / 1.10) / 2.3
                else:
                    coste_calculado = (pvp / 1.21) / 2
                
                return unidades * coste_calculado
            
            return 0
        except:
            return 0
    
    ventas_with_costs['Coste'] = ventas_with_costs.apply(calculate_cost, axis=1)
    
    # Calcular Beneficio
    def calculate_beneficio(row):
        try:
            importe = row['Importe'] if pd.notna(row['Importe']) else 0
            coste = row['Coste'] if pd.notna(row['Coste']) else 0
            iva = obtener_iva_articulo(row['Artículo'])
            beneficio = (importe / (1 + iva / 100)) - coste
            return beneficio
        except:
            return 0
    
    ventas_with_costs['Beneficio'] = ventas_with_costs.apply(calculate_beneficio, axis=1)
    
    # Seleccionar solo las columnas necesarias
    columnas_ventas = ['Vendedor', 'Serie', 'Documento', 'Fecha', 'Factura', 
                       'Artículo', 'Nombre artículo', 'Talla', 'Color', 
                       'Unidades', 'Precio', 'Importe', 'Comisión', 'Tipo registro',
                       'Coste', 'Beneficio']
    
    ventas_df = ventas_with_costs[columnas_ventas].copy()
    
    # Convertir columnas a tipos numéricos correctos
    ventas_df['Unidades'] = pd.to_numeric(ventas_df['Unidades'], errors='coerce').fillna(0)
    ventas_df['Importe'] = pd.to_numeric(ventas_df['Importe'], errors='coerce').fillna(0)
    ventas_df['Coste'] = pd.to_numeric(ventas_df['Coste'], errors='coerce').fillna(0)
    ventas_df['Beneficio'] = pd.to_numeric(ventas_df['Beneficio'], errors='coerce').fillna(0)
    
    # Resumen del procesamiento
    ventas_sin_coste = (ventas_with_costs['Coste'] == 0).sum()
    print(f"VENTAS procesadas: {len(ventas_df)} registros")
    print(f"  - Con coste encontrado: {len(ventas_df) - ventas_sin_coste}")
    print(f"  - Sin coste (asignado 0): {ventas_sin_coste}")
    print(f"\nTotal importe ventas: {ventas_df['Importe'].sum():.2f} €")
    print(f"Total coste ventas: {ventas_df['Coste'].sum():.2f} €")
    print(f"Total beneficio: {ventas_df['Beneficio'].sum():.2f} €")
    
    # Convertir fechas
    compras_df['Fecha'] = pd.to_datetime(compras_df['Fecha'])
    ventas_df['Fecha'] = pd.to_datetime(ventas_df['Fecha'])
    
    # =========================================================================
    # NORMALIZACIÓN DE DATOS
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 2: NORMALIZACIÓN DE DATOS")
    print("=" * 80)
    
    def normalizar_articulo(df):
        df = df.copy()
        
        def convertir_articulo(valor):
            if pd.isna(valor):
                return ''
            valor_str = str(valor)
            if valor_str.endswith('.0'):
                valor_str = valor_str[:-2]
            return valor_str
        
        df['codigo_str'] = df['Artículo'].apply(convertir_articulo)
        df['nombre_str'] = df['Nombre artículo'].astype(str).str.rstrip()
        df['talla_str'] = df['Talla'].fillna('').astype(str).str.strip()
        df['color_str'] = df['Color'].fillna('').astype(str).str.strip()
        return df
    
    ventas_df = normalizar_articulo(ventas_df)
    compras_df = normalizar_articulo(compras_df)
    stock_df = normalizar_articulo(stock_df)
    
    print("Columnas normalizadas creadas para comparación")
    
    # =========================================================================
    # FILTRAR ARTÍCULOS CON MENOS DE 10 DÍGITOS (REGLA PRIORITARIA)
    # =========================================================================
    
    # Esta regla tiene prioridad sobre todas las demás
    # Los artículos con códigos menores a 10 dígitos no se procesarán
    
    def codigo_valido(codigo):
        """Verifica que el código tenga al menos 10 dígitos"""
        if not codigo or codigo == 'nan':
            return False
        return len(codigo) >= 10
    
    compras_filas_antes = len(compras_df)
    ventas_filas_antes = len(ventas_df)
    stock_filas_antes = len(stock_df)
    
    # Filtrar artículos con códigos menores a 10 dígitos
    compras_df = compras_df[compras_df['codigo_str'].apply(codigo_valido)].copy()
    ventas_df = ventas_df[ventas_df['codigo_str'].apply(codigo_valido)].copy()
    stock_df = stock_df[stock_df['codigo_str'].apply(codigo_valido)].copy()
    
    print(f"\nFiltrados {compras_filas_antes - len(compras_df)} artículos con menos de 10 dígitos en COMPRAS")
    print(f"Filtrados {ventas_filas_antes - len(ventas_df)} artículos con menos de 10 dígitos en VENTAS")
    print(f"Filtrados {stock_filas_antes - len(stock_df)} artículos con menos de 10 dígitos en STOCK")
    
    # =========================================================================
    # FILTRAR FILAS CON UNIDADES = 0
    # =========================================================================
    
    compras_filas_antes = len(compras_df)
    ventas_filas_antes = len(ventas_df)
    stock_filas_antes = len(stock_df)
    
    compras_df = compras_df[compras_df['Unidades'].notna() & (compras_df['Unidades'] > 0)].copy()
    ventas_df = ventas_df[ventas_df['Unidades'].notna() & (ventas_df['Unidades'] > 0)].copy()
    stock_df = stock_df[stock_df['Unidades'].notna() & (stock_df['Unidades'] > 0)].copy()
    
    print(f"\nFiltradas {compras_filas_antes - len(compras_df)} filas con 0 unidades en COMPRAS")
    print(f"Filtradas {ventas_filas_antes - len(ventas_df)} filas con 0 unidades en VENTAS")
    print(f"Filtradas {stock_filas_antes - len(stock_df)} filas con 0 unidades en STOCK")
    
    # =========================================================================
    # PROCESAR SECCIONES
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("FASE 3: PROCESAMIENTO DE SECCIONES")
    print("=" * 80)
    
    # Determinar qué secciones procesar
    if seccion_especifica:
        secciones_a_procesar = [(seccion_especifica, SECCIONES[seccion_especifica])]
    else:
        secciones_a_procesar = list(SECCIONES.items())
    
    # Procesar cada sección
    estadisticas = []
    secciones_procesadas = []
    secciones_sin_datos = []
    
    for nombre_seccion, seccion_info in secciones_a_procesar:
        resultado = procesar_seccion(
            compras_df, ventas_df, stock_df, coste_df,
            nombre_seccion, seccion_info,
            FECHA_INICIO, FECHA_FIN, DIAS_PERIODO
        )
        
        if resultado:
            estadisticas.append(resultado)
            secciones_procesadas.append(nombre_seccion)
            print(f"\n✓ Sección '{nombre_seccion}' completada: {resultado['archivo']}")
        else:
            secciones_sin_datos.append(nombre_seccion)
    
    # =========================================================================
    # RESUMEN FINAL
    # =========================================================================
    
    print("\n" + "=" * 80)
    print("RESUMEN DEL PROCESAMIENTO")
    print("=" * 80)
    
    if args.periodo:
        periodo_info = DEFINICION_PERIODOS[args.periodo]
        print(f"\nPeríodo: {periodo_info['nombre']} ({periodo_info['descripcion']})")
        print(f"Fechas: {FECHA_INICIO.strftime('%d/%m/%Y')} - {FECHA_FIN.strftime('%d/%m/%Y')} ({DIAS_PERIODO} días)")
    else:
        print(f"\nPeríodo: {FECHA_INICIO.strftime('%d/%m/%Y')} - {FECHA_FIN.strftime('%d/%m/%Y')} ({DIAS_PERIODO} días)")
    
    print(f"\nSecciones procesadas: {len(secciones_procesadas)}")
    if secciones_procesadas:
        print("  - " + "\n  - ".join(sorted(secciones_procesadas)))
    
    if secciones_sin_datos:
        print(f"\nSecciones sin datos (saltadas): {len(secciones_sin_datos)}")
        print("  - " + "\n  - ".join(sorted(secciones_sin_datos)))
    
    if estadisticas:
        print(f"\nArchivos generados:")
        total_articulos = 0
        for stat in estadisticas:
            print(f"  - {stat['archivo']}: {stat['total_articulos']} artículos "
                  f"(A:{stat['categoria_a']}, B:{stat['categoria_b']}, "
                  f"C:{stat['categoria_c']}, D:{stat['categoria_d']})")
            total_articulos += stat['total_articulos']
        
        print(f"\nTotal artículos en todos los archivos: {total_articulos}")
    
    print("\n" + "=" * 80)
    print("PROCESO COMPLETADO CORRECTAMENTE")
    print("=" * 80)

# ============================================================================
# PUNTO DE ENTRADA
# ============================================================================

if __name__ == "__main__":
    main()